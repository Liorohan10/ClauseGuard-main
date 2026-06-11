from datetime import datetime, timezone
from io import BytesIO
from uuid import uuid4

from fastapi import APIRouter, Depends, HTTPException, Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from clauseguard.api.deps import get_es_service, get_openai_assistant
from clauseguard.models.openai_legal import ContractReviewOutput, FinalDecision, DecisionOutcome
from clauseguard.openai_assistant import OpenAILegalAssistant
from clauseguard.services.elasticsearch_service import ElasticsearchService

router = APIRouter(prefix="/review", tags=["review"])


def _build_review_record(contract_id: str, contract_filename: str, review: ContractReviewOutput) -> dict:
    reviewed_at = datetime.now(timezone.utc).isoformat()
    payload = review.model_dump(mode="json")

    # Count only meaningful findings (privacy + export control compliance findings)
    findings_count = len(review.compliance_findings) + len(review.missing_protections)

    return {
        "review_id": str(uuid4()),
        "contract_id": contract_id,
        "contract_filename": contract_filename,
        "reviewed_at": reviewed_at,
        "contract_safety_score": review.contract_safety_score,
        "summary": review.summary,
        "findings_count": findings_count,
        "export_control_triggered": review.export_control_triggered,
        "final_decision_outcome": (
            str(review.final_decision.outcome)
            if review.final_decision
            else "escalate"
        ),
        "review": payload,
    }


def _source_map_for_clauses(clauses: list[dict]) -> str:
    lines: list[str] = []
    for clause in clauses:
        excerpt = (clause.get("text", "") or "").strip()
        if len(excerpt) > 400:
            excerpt = excerpt[:400] + "..."
        lines.append(
            " | ".join(
                [
                    f"clause_id={clause.get('clause_id', '')}",
                    f"page={clause.get('page_number', '')}",
                    f"section={clause.get('section_number', '')}",
                    f"type={clause.get('clause_type', '')}",
                    f"offsets={clause.get('char_offset_start', '')}-{clause.get('char_offset_end', '')}",
                    f"excerpt={excerpt}",
                ]
            )
        )
    return "\n".join(lines)


def _review_from_record(record: dict) -> ContractReviewOutput:
    return ContractReviewOutput.model_validate(record.get("review") or {})


def _severity_rank(value: str) -> int:
    return {"critical": 5, "high": 4, "medium": 3, "low": 2, "info": 1}.get(str(value).lower(), 0)


def _citation_text(page: int | None, section: str, clause_id: str) -> str:
    parts = [
        f"p. {page}" if page else "",
        f"sec. {section}" if section else "",
        f"clause {clause_id}" if clause_id else "",
    ]
    return " · ".join(part for part in parts if part)


def _law_list(laws: list[str]) -> str:
    return "; ".join(laws) if laws else ""


def _decision_label(review: ContractReviewOutput) -> str:
    if review.final_decision and hasattr(review.final_decision, "outcome"):
        return str(review.final_decision.outcome).replace("_", " ").upper()
    return "ESCALATE"


def _export_rows(review: ContractReviewOutput) -> list[tuple]:
    """
    Build rows for the XLSX export.
    Columns: Issue ID | Domain | Clause/Section | Jurisdiction(s) | Finding/Risk | Risk Level |
             Applicable Laws | Recommended Redline | Fallback Position
    """
    rows: list[tuple] = []

    # Privacy and export-control compliance findings (spec sections 3 & 4)
    for item in review.compliance_findings:
        if str(item.status).lower() in ("pass", "not-applicable"):
            continue  # Only export actionable findings
        rows.append((
            item.issue_id or "",
            item.domain.upper() if hasattr(item, "domain") else "PRIVACY",
            _citation_text(item.source_page, item.source_section, item.source_clause_id) or item.requirement,
            ", ".join(item.jurisdictions) if item.jurisdictions else "",
            item.explanation,
            str(item.severity).upper(),
            _law_list(item.applicable_laws),
            item.remediation,
            item.fallback_position,
        ))

    # Missing protections
    for item in review.missing_protections:
        rows.append((
            "",
            (item.domain.upper() if hasattr(item, "domain") else "PRIVACY"),
            _citation_text(item.source_page, item.source_section, item.source_clause_id) or item.protection,
            "",
            item.why_missing,
            "HIGH",
            _law_list(item.applicable_laws if hasattr(item, "applicable_laws") else []),
            item.suggested_clause or item.mitigation,
            item.mitigation,
        ))

    # Proposed redlines (spec section 5)
    for item in review.redline_suggestions:
        rows.append((
            item.issue_id or "",
            item.domain.upper() if hasattr(item, "domain") else "PRIVACY",
            item.clause_reference,
            "",
            "Proposed Redline",
            "REDLINE",
            _law_list(item.applicable_laws),
            item.proposed_wording or item.drafting_instruction,
            item.drafting_instruction if item.proposed_wording else "",
        ))

    return rows


def _build_workbook(review: ContractReviewOutput, contract_filename: str, reviewed_at: str) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Compliance Report"

    # Colour palette
    title_fill = PatternFill("solid", fgColor="0F243E")
    header_fill = PatternFill("solid", fgColor="2F5D8A")
    light_fill = PatternFill("solid", fgColor="F3F6FA")
    green_fill = PatternFill("solid", fgColor="D9EAD3")
    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    red_fill = PatternFill("solid", fgColor="F4CCCC")
    critical_fill = PatternFill("solid", fgColor="8B0000")
    blue_fill = PatternFill("solid", fgColor="C9DAF8")

    # Title
    ws.merge_cells("A1:I1")
    ws["A1"] = "DATA PRIVACY & EXPORT CONTROL COMPLIANCE REVIEW REPORT"
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Metadata rows
    ws["A3"] = "Contract Filename"
    ws["B3"] = contract_filename
    ws["E3"] = "Compliance Score"
    ws["F3"] = f"{review.contract_safety_score}/100"
    ws["A4"] = "Audit Timestamp"
    ws["B4"] = reviewed_at
    ws["E4"] = "Final Decision"
    ws["F4"] = _decision_label(review)
    ws["A5"] = "Export Control"
    ws["B5"] = "Triggered" if review.export_control_triggered else "Not Triggered"

    for cell in ("A3", "A4", "A5", "E3", "E4"):
        ws[cell].font = Font(bold=True)
    ws["F3"].font = Font(bold=True, color="1F4E79")
    ws["F4"].font = Font(bold=True, color="C00000")

    # Section: Jurisdiction Profile
    ws["A7"] = "Applicable Jurisdictions"
    ws["A7"].font = Font(bold=True, size=12)

    row_idx = 8
    if review.jurisdiction_profile:
        privacy_juris = [j.jurisdiction for j in review.jurisdiction_profile.privacy_jurisdictions]
        ws[f"A{row_idx}"] = "Privacy Jurisdictions:"
        ws[f"B{row_idx}"] = ", ".join(privacy_juris) if privacy_juris else "None identified"
        ws[f"A{row_idx}"].font = Font(bold=True)
        row_idx += 1

        if review.export_control_triggered:
            export_juris = [j.jurisdiction for j in review.jurisdiction_profile.export_control_jurisdictions]
            ws[f"A{row_idx}"] = "Export Control Jurisdictions:"
            ws[f"B{row_idx}"] = ", ".join(export_juris) if export_juris else "See findings"
            ws[f"A{row_idx}"].font = Font(bold=True)
            row_idx += 1

    row_idx += 1

    # Findings table header
    ws[f"A{row_idx}"] = "Detailed Compliance Findings"
    ws[f"A{row_idx}"].font = Font(bold=True, size=12)
    row_idx += 1

    headers = [
        "Issue ID",
        "Domain",
        "Clause / Section",
        "Jurisdiction(s)",
        "Finding",
        "Risk Level",
        "Applicable Laws",
        "Recommended Redline / Mitigation",
        "Fallback Position",
    ]
    header_row = row_idx
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_start = header_row + 1
    for row_offset, row in enumerate(_export_rows(review), start=data_start):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_offset, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col_idx == 6:  # Risk Level column
                sev = str(value).lower()
                if sev == "critical":
                    cell.fill = critical_fill
                    cell.font = Font(color="FFFFFF", bold=True)
                elif _severity_rank(sev) >= 4:
                    cell.fill = red_fill
                elif _severity_rank(sev) == 3:
                    cell.fill = yellow_fill
                elif sev == "redline":
                    cell.fill = blue_fill
                else:
                    cell.fill = green_fill
            elif row_offset % 2 == 0:
                cell.fill = light_fill

    # Column widths
    widths = [12, 14, 30, 22, 52, 12, 30, 52, 40]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    ws.freeze_panes = f"A{data_start}"
    ws.sheet_view.showGridLines = False

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@router.post("/{contract_id}", response_model=ContractReviewOutput)
async def review_contract(
    contract_id: str,
    assistant: OpenAILegalAssistant = Depends(get_openai_assistant),
    es: ElasticsearchService = Depends(get_es_service),
):
    """Run Data Privacy & Export Control compliance review on a stored contract."""
    contract = await es.get_contract(contract_id)
    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")

    clauses = await es.get_clauses_by_contract(contract_id)
    if not clauses:
        raise HTTPException(status_code=404, detail="No clauses found for contract")

    ordered_clauses = sorted(
        clauses,
        key=lambda clause: (
            clause.get("page_number", 1),
            clause.get("char_offset_start", 0),
            clause.get("clause_type", ""),
        ),
    )
    contract_text = "\n\n".join(
        f"[{clause.get('clause_type', 'other')}] {clause.get('text', '').strip()}".strip()
        for clause in ordered_clauses
        if clause.get("text", "").strip()
    )

    if not contract_text.strip():
        raise HTTPException(status_code=422, detail="Stored contract text is empty")

    review = await assistant.analyze_contract_text(
        contract_text,
        filename=contract.get("filename", contract_id),
        source_context=_source_map_for_clauses(ordered_clauses),
    )
    review_record = _build_review_record(contract_id, contract.get("filename", contract_id), review)
    await es.save_review(review_record)
    await es.update_contract_review_summary(contract_id, review_record)
    return review


@router.get("/{contract_id}/latest", response_model=ContractReviewOutput)
async def latest_review(
    contract_id: str,
    es: ElasticsearchService = Depends(get_es_service),
):
    record = await es.get_latest_review(contract_id)
    if not record:
        raise HTTPException(status_code=404, detail="Review not found")
    return _review_from_record(record)


@router.get("/{contract_id}/history")
async def review_history(
    contract_id: str,
    es: ElasticsearchService = Depends(get_es_service),
):
    reviews = await es.list_reviews_by_contract(contract_id)
    return [
        {
            "review_id": review.get("review_id"),
            "reviewed_at": review.get("reviewed_at"),
            "contract_filename": review.get("contract_filename", contract_id),
            "contract_safety_score": review.get("contract_safety_score", 0),
            "summary": review.get("summary", ""),
            "findings_count": review.get("findings_count", 0),
            "export_control_triggered": review.get("export_control_triggered", False),
            "final_decision_outcome": review.get("final_decision_outcome", "escalate"),
        }
        for review in reviews
    ]


@router.get("/{contract_id}/export.xlsx")
async def export_review_xlsx(
    contract_id: str,
    es: ElasticsearchService = Depends(get_es_service),
):
    record = await es.get_latest_review(contract_id)
    if not record:
        raise HTTPException(status_code=404, detail="Review not found")

    review = _review_from_record(record)
    reviewed_at = record.get("reviewed_at", "")
    contract_filename = record.get("contract_filename", contract_id)
    buffer = _build_workbook(review, contract_filename, reviewed_at)
    filename = f"{contract_filename.rsplit('.', 1)[0]}_privacy_export_review.xlsx"
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{contract_id}/{review_id}", response_model=ContractReviewOutput)
async def get_review_by_id(
    contract_id: str,
    review_id: str,
    es: ElasticsearchService = Depends(get_es_service),
):
    record = await es.get_review_by_id(review_id)
    if not record or record.get("contract_id") != contract_id:
        raise HTTPException(status_code=404, detail="Review not found")
    return _review_from_record(record)


@router.delete("/{contract_id}")
async def delete_contract_bundle(
    contract_id: str,
    es: ElasticsearchService = Depends(get_es_service),
):
    await es.delete_reviews_by_contract(contract_id)
    await es.delete_clauses_by_contract(contract_id)
    await es.delete_contract(contract_id)
    return {"status": "deleted", "contract_id": contract_id}
